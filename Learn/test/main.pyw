import openpyxl
import sys
from PyQt5.QtWidgets import *
from PyQt5 import QtGui
from PyQt5.QtCore import QSize, Qt, QRect

# Наследуемся от QMainWindow
class MainWindow(QMainWindow):
    # Переопределяем конструктор класса
    def __init__(self):
        # Обязательно нужно вызвать метод супер класса
        QMainWindow.__init__(self)

        self.setMinimumSize(QSize(820, 500))            # Устанавливаем размеры
        self.setWindowTitle("Entrant data processing")  # Устанавливаем заголовок окна
        central_widget = QWidget(self)                  # Создаём центральный виджет
        self.setCentralWidget(central_widget)           # Устанавливаем центральный виджет
        self.setWindowIcon(QtGui.QIcon('icon.png'))     # Устанавливаем иконку приложения

        grid_layout = QGridLayout()             # Создаём сетку
        central_widget.setLayout(grid_layout)   # Устанавливаем данное размещение в центральный виджет

        global table
        table = QTableWidget(self) # Создаём таблицу
        table.setColumnCount(7)     # Устанавливаем три колонки
       # table.setRowCount(1)        # и одну строку в таблице

        # Устанавливаем заголовки таблицы
        table.setHorizontalHeaderLabels(["ID", "Name", "Surname", "Phone_number", "Avg_Math", "Avg_Phys", "Avg_Sertificate"])
        table.setSizeAdjustPolicy(QAbstractScrollArea.AdjustToContents)

        # Устанавливаем выравнивание на заголовки
        header = table.horizontalHeader()
        for i in range(7):
            table.horizontalHeaderItem(i).setTextAlignment(Qt.AlignHCenter)

        grid_layout.addWidget(table, 10, 0, 25, 25)   # Добавляем таблицу в сетку

        # groupboxes
        groupbox_exerpt = QGroupBox("Excertp")
        groupbox_exerpt.setAlignment(Qt.AlignHCenter)



        # подписи
        label7 = QLabel(groupbox_exerpt)
        label7.setText('Agv. Math from')
        label8 = QLabel(groupbox_exerpt)
        label8.setText('Avg. Phys from')
        label9 = QLabel(groupbox_exerpt)
        label9.setText('Avg. certificate from')
        label10 = QLabel(groupbox_exerpt)
        label10.setText('to')
        label11 = QLabel(groupbox_exerpt)
        label11.setText('to')
        label12 = QLabel(groupbox_exerpt)
        label12.setText('to')

        # группа обьектов выборки
        global spinbox1, spinbox2, spinbox3, spinbox11, spinbox22, spinbox33
        spinbox1 = QSpinBox(groupbox_exerpt)
        spinbox2 = QSpinBox(groupbox_exerpt)
        spinbox3 = QSpinBox(groupbox_exerpt)
        spinbox11 = QSpinBox(groupbox_exerpt)
        spinbox22 = QSpinBox(groupbox_exerpt)
        spinbox33 = QSpinBox(groupbox_exerpt)
        spinBoxList = [spinbox1, spinbox2, spinbox3]
        spinBoxList1 = [spinbox11, spinbox22, spinbox33]
        for i in spinBoxList1:
            i.setValue(5)
            i.setMinimum(0)
            i.setMaximum(5)
        for i in spinBoxList:
            i.setMinimum(0)
            i.setMaximum(5)

        # Добавленеи кнопок
        global okayButton, backButton
        okayButton = QPushButton('Okay', self) # кнопка 'Okay'
        okayButton.clicked.connect(self.dataChange) # привзяка метода 'dataChange' к кнопке
        backButton = QPushButton('Cancel', self)
        backButton.setEnabled(False)
        backButton.clicked.connect(self.backChange)

        #сетка table(0, 0, - 5, 1)
        grid_layout.addWidget(groupbox_exerpt, 0, 0, 6, 7)

        #сетка надписей
        grid_layout.addWidget(label7, 2, 1)
        grid_layout.addWidget(label8, 3, 1)
        grid_layout.addWidget(label9, 4, 1)
        grid_layout.addWidget(label10, 2, 3)
        grid_layout.addWidget(label11, 3, 3)
        grid_layout.addWidget(label12, 4, 3)
        #сетка spinbox
        grid_layout.addWidget(spinbox1, 2, 2)
        grid_layout.addWidget(spinbox2, 3, 2)
        grid_layout.addWidget(spinbox3, 4, 2)
        grid_layout.addWidget(spinbox11, 2, 4)
        grid_layout.addWidget(spinbox22, 3, 4)
        grid_layout.addWidget(spinbox33, 4, 4)
        #сетка кнопки
        grid_layout.addWidget(okayButton, 3, 5)
        grid_layout.addWidget(backButton, 4, 5)


        # Добавление тулбокса
        openFile = QAction('Open', self)
        openFile.setShortcut('Ctrl+O')
        openFile.setStatusTip('Open new File')
        openFile.triggered.connect(self.openFile)

        saveFile = QAction('Save as', self)
        saveFile.setShortcut('Ctrl+S')
        saveFile.setStatusTip('Save file As')
        saveFile.triggered.connect(self.saveFileAs)


        exitAction = QAction('Exit', self)
        exitAction.setShortcut('Ctrl+Q')
        exitAction.setStatusTip('Exit application')
        exitAction.triggered.connect(self.close)
        self.statusBar()

        aboutAction = QAction('About', self)
        aboutAction.triggered.connect(self.aboutShow)

        menubar = self.menuBar()
        fileMenu = menubar.addMenu('&File')
        fileMenu.addAction(openFile)
        fileMenu.addAction(saveFile)
        fileMenu.addAction(exitAction)
        aboutMenu = menubar.addMenu('&About')
        aboutMenu.addAction(aboutAction)

    def aboutShow(self): # метод вызова окна 'About'
        aboutShow = QMessageBox()
        aboutShow.setIcon(QMessageBox.Information) # установка иконки окна
        aboutShow.setWindowTitle('About') # установка названия окна
        aboutShow.setText('Application developed by student Yaroslav Reshetnyk, group: IPZ/i-17k') # установка текста окна
        aboutShow.addButton('Okay', QMessageBox.AcceptRole) # установка кнопки 'Okey'
        aboutShow.exec() # запуск окна

    def openFile(self): # метод открытие файла
        try:
            table.setSortingEnabled(False)
            table.setRowCount(0) # очистка рабочего листа перед загрузкой для избежания ошибок
            openFileName = QFileDialog.getOpenFileName(self, 'Open file', 'C:\\', "Excel Workbook (*.xlsx)")[0]
            importTable = openpyxl.load_workbook(openFileName) # загрузка в openpyxl файла
            firstElement = importTable.sheetnames # получаем список листов
            sheetsHappens = importTable[firstElement[0]] # получаем название рабоче листа
            table.setRowCount(sheetsHappens.max_row) # устанавливаем максимальное количество строк
            listOfBukvi = ['A', 'B', 'C', 'D', 'E', 'F', 'G'] # список используемых колонок в экселе
            for i in range(1, sheetsHappens.max_row+1): # цикл заполнения таблицы
                for j in range(1, 8):
                    x = sheetsHappens[listOfBukvi[j-1] + str(i)].value
                    if x != None:
                        table.setItem(i - 1, j - 1, QTableWidgetItem(str(x)))
                    else:
                        continue
            table.resizeColumnsToContents()
            self.strToFloat()
        except:
            pass # в случае закрытия диалогового окна открытия - бездействие

    def strToFloat(self): # пробразование стринговых значений оценок в float
        table.setSortingEnabled(True)
        for i in range(table.rowCount()): # столбец ID
            item = table.item(i, 0).text()
            cellitem = QTableWidgetItem()
            cellitem.setData(Qt.DisplayRole, int(item))
            table.setItem(i, 0, cellitem)
        for i in range(table.rowCount()): # столбец Math
            item = table.item(i, 4).text()
            cellitem = QTableWidgetItem()
            cellitem.setData(Qt.DisplayRole, float(item))
            table.setItem(i, 4, cellitem)
        for i in range(table.rowCount()): # столбец Phys
            item = table.item(i, 5).text()
            cellitem = QTableWidgetItem()
            cellitem.setData(Qt.DisplayRole, float(item))
            table.setItem(i, 5, cellitem)
        for i in range(table.rowCount()): # столбец AVG
            item = table.item(i, 6).text()
            cellitem = QTableWidgetItem()
            cellitem.setData(Qt.DisplayRole, float(item))
            table.setItem(i, 6, cellitem)

    def saveFileAs(self): # метод сохранения файла
        try:
            saveFileName = QFileDialog.getSaveFileName(self, 'Save File', 'С:\\', "Excel Worbook (*.xlsx)")[0] #вызываем диалоговое окно с выбором места сохранения
            saveBook = openpyxl.Workbook() # открываем рабочий лист
            saveSheet = saveBook.active # активируем рабочий лист
            for i in range(table.rowCount()):
                for j in range(table.columnCount()):
                    try:
                        cell = saveSheet.cell(row=i+1, column=j+1) # определяем номер столбца и строки внутри файла
                        item = table.item(i, j).text() # получаем данные с ячеек таблицы программы
                        cell.value = item # записываем данные в файл
                    except:
                        item = ' ' # на случай пустой ячейки
            saveBook.save(saveFileName) # сохраняем файл под выбранным название и в выбранной директории
        except:
            pass # в случае закрытия диалогового окна сохранения - бездействие

    def dataChange(self): # метод создания выборки по трём условиям
        table.setSortingEnabled(False) # выключаем режим сортировки во избежания конфликтов
        okayButton.setEnabled(False) # выключаем кнопку 'Okay'
        backButton.setEnabled(True) # включаем кнопку 'Cancel'
        global dataList
        dataList = []
        filterList = []
        for i in range(table.rowCount()): # сохраняем перед созданием выборки данные таблицы в двумерную матрицу
            listData = []
            for j in range(table.columnCount()):
                try:
                    item = table.item(i, j).text()
                    listData.append(item)
                except:
                    listData.append('')
            dataList.append(listData)
        table.setRowCount(0) # устанавливаем количество строк таблицы '0'
        rows = 0 # количество строк подошедших под условия
        for i in dataList: # цикл создания выборки за условиями и дальнейшего сохранения в одномерный масив
            if float(i[4]) >= spinbox1.value() and float(i[4]) <= spinbox11.value() \
                    and float(i[5]) >= spinbox2.value() and float(i[5]) <= spinbox22.value() \
                    and float(i[6]) >= spinbox3.value() and float(i[6]) <= spinbox33.value():
                for j in i:
                    filterList.append(j)
                rows += 1
        table.setRowCount(rows) # устанавливаем количество строк обновлённой таблицы с выборкой
        for i in range(rows): # цикл заполнения таблицы значениями
            for j in range(7):
                x = filterList.pop(0)
                table.setItem(i, j, QTableWidgetItem(str(x)))
        self.strToFloat() # вызов метода strToFloat

    def backChange(self): # метод возврата после выборки
        table.setSortingEnabled(False) # выключаем режим сортировки во избежания конфликтов
        okayButton.setEnabled(True) # включаем кнопку 'Okay'
        backButton.setEnabled(False) # выключаем кнопку 'Cancel'
        table.setRowCount(0) # устанавливаем количество строк таблицы '0'
        backlist = []
        backrows = 0 # количество строк подошедших под условия
        for i in dataList: # преобразование двумерной масива в одномерный масив
            for j in i:
                backlist.append(j)
            backrows += 1
        table.setRowCount(backrows)
        for i in range(backrows): # цикл заполнения таблицы значениями
            for j in range(7):
                x = backlist.pop(0)
                table.setItem(i, j, QTableWidgetItem(str(x)))
        self.strToFloat()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    mw = MainWindow()
    mw.show()
    sys.exit(app.exec())